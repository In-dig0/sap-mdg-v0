"""
MDG — Migration Data Governance
POSIZIONE: mdg-v0/streamlit/app/pages/10_access_log.py

Log Accessi — visualizzazione eventi login/logout (solo admin_role).
"""

import os
import io
from datetime import date
import streamlit as st
import pandas as pd
import requests

from mdg_auth import require_role, render_sidebar_menu

st.set_page_config(
    page_title="Log Accessi",
    page_icon="🔐",
    layout="wide",
)

require_role("admin_role")
render_sidebar_menu()

PIPELINE_API_URL = os.getenv("PIPELINE_API_URL", "http://mdg_fastapi:8000")


def fetch_log(limit: int = 200, email: str = "", success_filter: str = "Tutti") -> pd.DataFrame:
    params = {"limit": limit}
    if email:
        params["email"] = email
    if success_filter == "Solo OK":
        params["success"] = "true"
    elif success_filter == "Solo Falliti":
        params["success"] = "false"

    token = st.session_state.get("mdg_token", "")
    try:
        r = requests.get(
            f"{PIPELINE_API_URL}/auth/access-log",
            params=params,
            headers={"Authorization": f"Bearer {token}"},
            timeout=10,
        )
        r.raise_for_status()
        data = r.json()
        if not data:
            return pd.DataFrame()
        df = pd.DataFrame(data)
        df["logged_at"] = pd.to_datetime(df["logged_at"]).dt.tz_localize(None)
        df["logged_at"] = df["logged_at"].dt.strftime("%d/%m/%Y %H:%M:%S")
        return df
    except Exception as e:
        st.error(f"Errore nel recupero log: {e}")
        return pd.DataFrame()


def df_to_xlsx(df: pd.DataFrame) -> bytes:
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Log Accessi")
        ws = writer.sheets["Log Accessi"]
        border_side   = Side(style="thin", color="000000")
        header_border = Border(left=border_side, right=border_side,
                               top=border_side, bottom=border_side)
        for cell in ws[1]:
            cell.font      = Font(bold=True)
            cell.fill      = PatternFill("solid", start_color="BFBFBF", end_color="BFBFBF")
            cell.border    = header_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.auto_filter.ref = ws.dimensions
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Header
# ---------------------------------------------------------------------------
st.markdown(
    '<h1 style="color:#38BDF8;">🔐 MDG — Log Accessi</h1>',
    unsafe_allow_html=True,
)
st.caption(":yellow[Storico degli eventi di login all'applicazione. Visibile solo agli amministratori.]")
st.divider()

# ---------------------------------------------------------------------------
# Filtri
# ---------------------------------------------------------------------------
col_email, col_esito, col_limit, _ = st.columns([3, 2, 2, 3])

with col_email:
    filter_email = st.text_input("Filtra per email", placeholder="es. mario.rossi@...")

with col_esito:
    filter_esito = st.selectbox(
        "Esito",
        options=["Tutti", "Solo OK", "Solo Falliti"],
    )

with col_limit:
    filter_limit = st.selectbox(
        "Max righe",
        options=[100, 200, 500, 1000],
        index=1,
    )

# ---------------------------------------------------------------------------
# Caricamento dati
# ---------------------------------------------------------------------------
df = fetch_log(limit=filter_limit, email=filter_email, success_filter=filter_esito)

# ---------------------------------------------------------------------------
# KPI
# ---------------------------------------------------------------------------
if not df.empty:
    n_total   = len(df)
    n_ok      = int((df["success"] == True).sum())
    n_fail    = int((df["success"] == False).sum())
    pct_fail  = round(n_fail / n_total * 100, 1) if n_total > 0 else 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Eventi totali",   n_total)
    c2.metric("✅ Login OK",     n_ok)
    c3.metric("❌ Login Falliti", n_fail)
    c4.metric("% Falliti",       f"{pct_fail}%")
    st.divider()

# ---------------------------------------------------------------------------
# Tabella
# ---------------------------------------------------------------------------
if df.empty:
    st.info("Nessun evento trovato con i filtri selezionati.")
else:
    # Rinomina e riordina colonne per la visualizzazione
    df_display = df[["logged_at", "email", "success", "role", "reason"]].copy()
    df_display.columns = ["Data/ora", "Email", "Esito", "Ruolo", "Motivo errore"]
    df_display["Esito"] = df_display["Esito"].map({True: "✅ OK", False: "❌ Fallito"})
    df_display["Motivo errore"] = df_display["Motivo errore"].fillna("—")
    df_display["Ruolo"] = df_display["Ruolo"].fillna("—")

    def highlight_row(row):
        if row["Esito"] == "❌ Fallito":
            return ["background-color: rgba(226,75,74,0.12)"] * len(row)
        return ["background-color: rgba(99,153,34,0.08)"] * len(row)

    styled = df_display.style.apply(highlight_row, axis=1)

    st.subheader(f"Ultimi {len(df_display)} eventi")
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        column_config={
            "Data/ora":      st.column_config.TextColumn(width="medium"),
            "Email":         st.column_config.TextColumn(width="medium"),
            "Esito":         st.column_config.TextColumn(width="small"),
            "Ruolo":         st.column_config.TextColumn(width="small"),
            "Motivo errore": st.column_config.TextColumn(width="large"),
        }
    )

    # Export Excel
    xlsx_data = df_to_xlsx(df_display)
    st.download_button(
        label="⬇️ Esporta in Excel",
        data=xlsx_data,
        file_name=f"access_log_{date.today().strftime('%Y%m%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
